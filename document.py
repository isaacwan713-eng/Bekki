import os
import re
from collections import Counter
import csv
from io import StringIO

SUPPORTED_EXTENSIONS = {
    ".txt",
    ".md",
    ".pdf",
    ".docx",
    ".csv",
    ".xlsx",
}


def read_document(file_path):
    """
    Read a local document and return normalized text.
    """

    if not file_path:
        return {
            "success": False,
            "text": "",
            "file_name": "",
            "file_path": "",
            "error": "No file path provided.",
        }

    if not os.path.exists(file_path):
        return {
            "success": False,
            "text": "",
            "file_name": os.path.basename(file_path),
            "file_path": file_path,
            "error": "File does not exist.",
        }

    extension = os.path.splitext(
        file_path
    )[1].lower()

    if extension not in SUPPORTED_EXTENSIONS:
        return {
            "success": False,
            "text": "",
            "file_name": os.path.basename(file_path),
            "file_path": file_path,
            "error": (
                "Unsupported file type: "
                + extension
            ),
        }

    try:
        if extension in {".txt", ".md"}:
            text = _read_text_file(file_path)

        elif extension == ".pdf":
            text = _read_pdf(file_path)

        elif extension == ".docx":
            text = _read_docx(file_path)

        elif extension == ".csv":
            text = _read_csv(file_path)

        elif extension == ".xlsx":
            text = _read_xlsx(file_path)     

        else:
            text = ""

    except Exception as error:
        return {
            "success": False,
            "text": "",
            "file_name": os.path.basename(file_path),
            "file_path": file_path,
            "error": str(error),
        }

    text = _clean_text(text)

    return {
        "success": True,
        "text": text,
        "file_name": os.path.basename(file_path),
        "file_path": file_path,
        "extension": extension,
        "length": len(text),
        "error": None,
    }


def _read_text_file(file_path):
    encodings = [
        "utf-8",
        "utf-8-sig",
        "gb18030",
        "latin-1",
    ]

    for encoding in encodings:
        try:
            with open(
                file_path,
                "r",
                encoding=encoding,
            ) as file:
                return file.read()

        except UnicodeDecodeError:
            continue

    raise ValueError(
        "Unable to decode text file."
    )


def _read_pdf(file_path):
    from pypdf import PdfReader

    reader = PdfReader(file_path)

    pages = []

    for page_number, page in enumerate(
        reader.pages,
        start=1,
    ):
        text = page.extract_text() or ""

        if text.strip():
            pages.append(
                f"[Page {page_number}]\n{text}"
            )

    return "\n\n".join(pages)

def _read_csv(file_path):
    raw_text = _read_text_file(file_path)

    if not raw_text.strip():
        return ""

    sample = raw_text[:8192]

    try:
        dialect = csv.Sniffer().sniff(
            sample,
            delimiters=",;\t|",
        )
    except csv.Error:
        dialect = csv.excel

    try:
        has_header = csv.Sniffer().has_header(
            sample
        )
    except csv.Error:
        has_header = False

    reader = csv.reader(
        StringIO(raw_text),
        dialect,
    )

    rows = list(reader)

    if not rows:
        return ""

    max_columns = max(
        len(row)
        for row in rows
    )

    if has_header:
        headers = [
            value.strip() or f"Column {index + 1}"
            for index, value in enumerate(rows[0])
        ]

        data_rows = rows[1:]
        first_row_number = 2

    else:
        headers = [
            f"Column {index + 1}"
            for index in range(max_columns)
        ]

        data_rows = rows
        first_row_number = 1

    while len(headers) < max_columns:
        headers.append(
            f"Column {len(headers) + 1}"
        )

    parts = [
        "[CSV Document]",
        "Columns: " + " | ".join(headers),
    ]

    for row_number, row in enumerate(
        data_rows,
        start=first_row_number,
    ):
        values = []

        for column_index, value in enumerate(row):
            value = value.strip()

            if not value:
                continue

            values.append(
                headers[column_index]
                + ": "
                + value
            )

        if values:
            parts.append(
                "\n[Row "
                + str(row_number)
                + "]\n"
                + "\n".join(values)
            )

    return "\n".join(parts)

def _format_spreadsheet_value(value):
    if value is None:
        return ""

    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except (TypeError, ValueError):
            pass

    return str(value).strip()


def _read_xlsx(file_path):
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(
        file_path,
        read_only=True,
        data_only=True,
    )

    try:
        parts = [
            "[Excel Workbook]",
            "File: " + os.path.basename(file_path),
        ]

        readable_sheet_count = 0

        for worksheet in workbook.worksheets:
            populated_rows = []

            for row_number, row in enumerate(
                worksheet.iter_rows(
                    values_only=True
                ),
                start=1,
            ):
                values = [
                    _format_spreadsheet_value(
                        value
                    )
                    for value in row
                ]

                while (
                    values
                    and not values[-1]
                ):
                    values.pop()

                if any(values):
                    populated_rows.append(
                        (
                            row_number,
                            values,
                        )
                    )

            if not populated_rows:
                continue

            readable_sheet_count += 1

            header_row_number, header_values = (
                populated_rows[0]
            )

            max_columns = max(
                len(values)
                for _, values in populated_rows
            )

            headers = []

            for column_index in range(
                max_columns
            ):
                if column_index < len(
                    header_values
                ):
                    header = (
                        header_values[
                            column_index
                        ]
                    )
                else:
                    header = ""

                if not header:
                    header = (
                        "Column "
                        + get_column_letter(
                            column_index + 1
                        )
                    )

                headers.append(header)

            parts.append(
                "\n[Sheet: "
                + worksheet.title
                + "]"
            )

            parts.append(
                "Header row: "
                + str(header_row_number)
            )

            parts.append(
                "Columns: "
                + " | ".join(headers)
            )

            for row_number, values in (
                populated_rows[1:]
            ):
                row_parts = []

                for column_index, value in enumerate(
                    values
                ):
                    if not value:
                        continue

                    row_parts.append(
                        headers[column_index]
                        + ": "
                        + value
                    )

                if row_parts:
                    parts.append(
                        "\n[Sheet: "
                        + worksheet.title
                        + " | Row "
                        + str(row_number)
                        + "]\n"
                        + "\n".join(
                            row_parts
                        )
                    )

        if readable_sheet_count == 0:
            return ""

        return "\n".join(parts)

    finally:
        workbook.close()


def _read_docx(file_path):
    from docx import Document

    document = Document(file_path)

    paragraphs = []

    for paragraph in document.paragraphs:
        text = paragraph.text.strip()

        if text:
            paragraphs.append(text)

    return "\n\n".join(paragraphs)




def _clean_text(text):
    if not text:
        return ""

    text = text.replace(
        "\x00",
        ""
    )

    lines = [
        line.rstrip()
        for line in text.splitlines()
    ]

    return "\n".join(lines).strip()

def chunk_document(
    text,
    chunk_size=2500,
    overlap=300,
):
    if not text:
        return []

    if chunk_size <= 0:
        raise ValueError(
            "chunk_size must be greater than 0."
        )

    if overlap < 0:
        raise ValueError(
            "overlap cannot be negative."
        )

    if overlap >= chunk_size:
        raise ValueError(
            "overlap must be smaller than chunk_size."
        )

    chunks = []

    start = 0
    chunk_index = 0
    text_length = len(text)

    while start < text_length:
        end = min(
            start + chunk_size,
            text_length
        )

        chunk_text = text[
            start:end
        ].strip()

        if chunk_text:
            chunks.append(
                {
                    "index": chunk_index,
                    "text": chunk_text,
                    "start": start,
                    "end": end,
                    "length": len(chunk_text),
                }
            )

            chunk_index += 1

        if end >= text_length:
            break

        start = end - overlap

    return chunks

def _tokenize(text):
    if not text:
        return []

    text = text.lower()

    # English words / numbers + individual CJK characters
    return re.findall(
        r"[a-z0-9]+|[\u4e00-\u9fff]",
        text
    )


def shortlist_chunks(
    query,
    chunks,
    limit=10,
):
    if not query or not chunks:
        return []

    query_tokens = _tokenize(query)

    if not query_tokens:
        return chunks[:limit]

    query_counter = Counter(
        query_tokens
    )

    scored_chunks = []

    for chunk in chunks:
        chunk_text = chunk.get(
            "text",
            ""
        )

        chunk_tokens = _tokenize(
            chunk_text
        )

        chunk_counter = Counter(
            chunk_tokens
        )

        score = 0

        for token, query_count in (
            query_counter.items()
        ):
            score += min(
                query_count,
                chunk_counter.get(
                    token,
                    0
                )
            )

        scored_chunks.append(
            {
                **chunk,
                "keyword_score": score,
            }
        )

    scored_chunks.sort(
        key=lambda item: item[
            "keyword_score"
        ],
        reverse=True,
    )

    # If nothing matched, still give the AI
    # some document content to inspect.
    matched = [
        item
        for item in scored_chunks
        if item["keyword_score"] > 0
    ]

    if matched:
        return matched[:limit]

    return scored_chunks[:limit]

def rerank_chunks(
    query,
    candidate_chunks,
    top_k=4,
):
    if not candidate_chunks:
        return []

    from tools import run_ai_prompt

    chunk_text = ""

    for chunk in candidate_chunks:
        chunk_text += (
            "\n\n"
            + "===== CHUNK "
            + str(chunk["index"])
            + " =====\n"
            + chunk["text"]
        )

    input_text = (
        "User question:\n"
        + query
        + "\n\nCandidate document chunks:"
        + chunk_text
    )

    result = run_ai_prompt(
        "prompts/document_rank.txt",
        input_text,
        expect_json=True,
        num_ctx=8192,
        num_predict=256,
    )

    if not isinstance(result, dict):
        return []

    ranked_indices = result.get(
        "ranked_indices",
        []
    )

    if not isinstance(
        ranked_indices,
        list
    ):
        return []

    chunk_map = {
        chunk["index"]: chunk
        for chunk in candidate_chunks
    }

    selected = []

    for index in ranked_indices:
        if index in chunk_map:
            selected.append(
                chunk_map[index]
            )

        if len(selected) >= top_k:
            break

    return selected

def retrieve_document_chunks(
    query,
    chunks,
    shortlist_limit=10,
    top_k=4,
):
    candidates = shortlist_chunks(
        query,
        chunks,
        limit=shortlist_limit,
    )

    print(
        "[DOCUMENT SHORTLIST]",
        [
            (
                item["index"],
                item.get(
                    "keyword_score",
                    0
                ),
            )
            for item in candidates
        ]
    )

    selected = rerank_chunks(
        query,
        candidates,
        top_k=top_k,
    )

    print(
        "[DOCUMENT SELECTED]",
        [
            item["index"]
            for item in selected
        ]
    )

    return selected

def build_document_context(
    file_name,
    selected_chunks
):
    if not selected_chunks:
        return (
            "Current Document:\n"
            + file_name
            + "\n\nNo relevant document evidence was found."
        )

    parts = [
        "Current Document:",
        file_name,
        "",
        "Relevant Document Evidence:",
    ]

    for chunk in selected_chunks:
        parts.append(
            "\n===== CHUNK "
            + str(chunk["index"])
            + " =====\n"
            + chunk["text"]
        )

    return "\n".join(parts)

_current_document = {
    "file_name": None,
    "file_path": None,
    "chunks": [],
}


def load_document(file_path):
    global _current_document

    result = read_document(
        file_path
    )

    if not result.get("success"):
        return result

    chunks = chunk_document(
        result["text"]
    )

    if not chunks:
        return {
            "success": False,
            "file_name": result["file_name"],
            "file_path": result["file_path"],
            "chunk_count": 0,
            "length": result.get("length", 0),
            "error": (
                "No readable text was found "
                "in this document."
            ),
        }

    _current_document = {
        "file_name": result["file_name"],
        "file_path": result["file_path"],
        "chunks": chunks,
    }

    return {
        "success": True,
        "file_name": result["file_name"],
        "file_path": result["file_path"],
        "chunk_count": len(chunks),
        "length": result["length"],
        "error": None,
    }


def has_document():
    return bool(
        _current_document.get("chunks")
    )


def get_current_document():
    return _current_document


def clear_document():
    global _current_document

    _current_document = {
        "file_name": None,
        "file_path": None,
        "chunks": [],
    }

def get_document_overview_context(
    document,
    max_chunks=6,
):
    chunks = document.get(
        "chunks",
        []
    )

    if not chunks:
        return ""

    if len(chunks) <= max_chunks:
        selected_chunks = chunks

    else:
        positions = [
            round(
                i * (len(chunks) - 1)
                / (max_chunks - 1)
            )
            for i in range(max_chunks)
        ]

        selected_chunks = [
            chunks[index]
            for index in positions
        ]

    print(
        "[DOCUMENT OVERVIEW CHUNKS]",
        [
            chunk["index"]
            for chunk in selected_chunks
        ]
    )

    return build_document_context(
        document["file_name"],
        selected_chunks,
    )

def get_document_context(query):
    if not has_document():
        return ""

    # Local import avoids circular import:
    # tools.py already imports document.py
    import tools

    current_document = get_current_document()

    mode = tools.decide_document_mode(
        query
    )

    if mode == "overview":
        return get_document_overview_context(
            current_document
        )

    selected_chunks = retrieve_document_chunks(
        query,
        current_document["chunks"],
    )

    return build_document_context(
        current_document["file_name"],
        selected_chunks,
    )